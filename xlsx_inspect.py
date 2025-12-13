from openpyxl import load_workbook
wb = load_workbook('sample_from_response.xlsx')
print(wb.sheetnames)
for ws in wb.worksheets:
    print('Sheet', ws.title)
    for row in ws.iter_rows(values_only=True):
        if any(row):
            print(row)
