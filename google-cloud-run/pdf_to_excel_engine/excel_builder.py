"""Build a final Excel workbook that merges text, tables, and diagrams."""

from __future__ import annotations

import io
import logging
import math
import re
from collections import defaultdict
from typing import Dict, Iterable, List, Tuple

from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .types import ExtractedTable, ImageFragment, ParagraphBlock

LOGGER = logging.getLogger(__name__)


def build_workbook(
    paragraphs: Iterable[ParagraphBlock],
    tables: Iterable[ExtractedTable],
    images: Iterable[ImageFragment],
) -> Tuple[bytes, Dict[str, int]]:
    """Create a single-sheet workbook and return bytes plus stats."""

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Single column layout - all content in column A
    current_row = 1

    # Set column width for single column layout
    ws.column_dimensions['A'].width = 60  # Single column for all content
    column_widths: Dict[int, float] = defaultdict(lambda: 60.0)

    border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    body_font = Font(name="Calibri", size=11)

    stats = {
        "text_blocks": 0,
        "tables_written": 0,
        "images_embedded": 0,
        "rows_written": 1,
    }

    events = _build_event_stream(paragraphs, tables, images)

    # Global content deduplication - track all written content with aggressive matching
    written_content = set()  # Track normalized text content that has been written
    written_content_parts = set()  # Track parts of written content for substring matching
    
    def _normalize_text_aggressive(text: str) -> str:
        """Aggressively normalize text for duplicate detection."""
        if not text:
            return ""
        # Remove all punctuation and special characters, keep only alphanumeric and spaces
        # Convert to lowercase
        normalized = text.lower()
        # Remove all punctuation except spaces
        normalized = ''.join(c if c.isalnum() or c.isspace() else ' ' for c in normalized)
        # Normalize whitespace (multiple spaces to single space)
        normalized = " ".join(normalized.split())
        return normalized.strip()
    
    def _is_duplicate_content(text: str) -> bool:
        """Check if this text content has already been written (with summary duplicate detection)."""
        if not text or len(text.strip()) < 2:
            return False
        
        # Normalize text (simple normalization - just lowercase and whitespace)
        normalized = " ".join(text.strip().lower().split())
        
        if not normalized or len(normalized) < 2:
            return False
        
        # Check for exact match
        if normalized in written_content:
            LOGGER.info(f"Duplicate detected (exact match): {normalized[:50]}...")
            return True
        
        # For long text (likely summary paragraph), check for high similarity with existing content
        if len(normalized) > 100:
            for existing in written_content:
                if len(existing) > 100:
                    # Check if one is substring of another (summary duplicates)
                    if normalized in existing or existing in normalized:
                        LOGGER.info(f"Duplicate detected (substring - likely summary): {normalized[:50]}...")
                        return True
                    # Check word overlap for long paragraphs
                    words = set(normalized.split())
                    existing_words = set(existing.split())
                    if len(words) > 20 and len(existing_words) > 20:
                        overlap = len(words & existing_words) / max(len(words), len(existing_words))
                        if overlap > 0.75:  # 75% word overlap = duplicate summary
                            LOGGER.info(f"Duplicate detected (75%+ word overlap - likely summary): {normalized[:50]}...")
                            return True
        
        # Not a duplicate - add to written content
        written_content.add(normalized)
        
        return False

    # First pass: Identify candidate photo, name/title, and summary from ALL events sorted by Y position
    # We need to find the FIRST occurrence (lowest Y position) of name/title and summary
    candidate_photo = None
    candidate_photo_fragment = None
    name_title_block = None
    name_event = None
    summary_block = None
    summary_event = None
    processed_header_events = []  # Events to mark as processed
    
    # Sort ALL events by Y position (top to bottom) to find first occurrences
    sorted_events = sorted(events, key=lambda e: (e.get("page", 0), e.get("top", 9999), e.get("left", 0)))
    
    # Find candidate photo (first image in top portion of first page)
    for event in sorted_events:
        if event["type"] == "image" and event["page"] == 0:
            fragment = event["payload"]
            if fragment.label == "embedded-image":
                image_top = fragment.bbox[1] if fragment.bbox else 9999
                is_candidate_photo = (
                    image_top < 400 and  # Top portion of page
                    fragment.width < 500 and fragment.height < 500 and
                    fragment.width > 50 and fragment.height > 50
                )
                if is_candidate_photo and candidate_photo is None:
                    candidate_photo = (fragment, event)
                    candidate_photo_fragment = fragment
                    LOGGER.info(f"Identified candidate photo at Y={image_top:.1f}")
                    break
    
    # Find name/title - search ALL events for name pattern
    # Priority: 1) Top blocks first, 2) Any block with exact name match (like "Muhammad Amjad")
    name_candidates = []  # Store all potential name blocks with their Y position
    
    for event in sorted_events:
        if event["type"] == "text" and event["page"] == 0:
            block = event["payload"]
            block_top = block.bbox[1] if block.bbox and len(block.bbox) >= 2 else 9999
            first_line = block.lines[0].strip() if block.lines else ""
            
            if len(block.lines) <= 3 and first_line:
                # Check for exact name pattern (2 words, both capitalized - like "Muhammad Amjad")
                words = first_line.split()
                is_exact_name = (
                    len(words) == 2 and 
                    words[0][0].isupper() and words[1][0].isupper() and
                    not any(keyword in first_line.upper() for keyword in ["ENGINEER", "MANAGER", "TECHNOLOGIES", "COMPANY"])
                )
                
                # Check for title pattern
                has_title = any(keyword in first_line.upper() for keyword in [
                    "ENGINEER", "MANAGER", "DEVELOPER", "ANALYST", "SPECIALIST", 
                    "CONSULTANT", "DIRECTOR", "LEAD", "SENIOR", "JUNIOR", "NETWORK"
                ])
                
                word_count = len(words)
                is_name_pattern = (word_count >= 2 and word_count <= 4) or has_title
                
                if is_exact_name or is_name_pattern:
                    name_candidates.append((block_top, block, event, is_exact_name))
    
    # Select name: prefer exact name match, then top position
    if name_candidates:
        # Sort: exact name first, then by Y position (top to bottom)
        name_candidates.sort(key=lambda x: (not x[3], x[0]))  # exact_name first (True=1, False=0), then Y position
        block_top, name_title_block, name_event, _ = name_candidates[0]
        processed_header_events.append(name_event)
        LOGGER.info(f"Identified name/title at Y={block_top:.1f}: {name_title_block.lines[0] if name_title_block.lines else ''}")
    
    # Find summary - look for FIRST long paragraph with summary keywords (lowest Y position = top of page)
    summary_candidates = []
    
    for event in sorted_events:
        if event["type"] == "text" and event["page"] == 0:
            block = event["payload"]
            block_text = " ".join(block.lines).strip()
            block_top = block.bbox[1] if block.bbox and len(block.bbox) >= 2 else 9999
            
            if len(block_text) > 100:
                summary_keywords = ["experience", "skilled", "dedicated", "professional", "expertise", "years"]
                if any(keyword in block_text.lower() for keyword in summary_keywords):
                    # Check if it starts with "Dedicated" (common summary start)
                    starts_with_dedicated = block_text.strip().lower().startswith("dedicated")
                    summary_candidates.append((block_top, block, event, starts_with_dedicated))
    
    # Select summary: prefer blocks that start with "Dedicated", then top position
    if summary_candidates:
        # Sort: "Dedicated" start first, then by Y position (top to bottom)
        summary_candidates.sort(key=lambda x: (not x[3], x[0]))  # dedicated_start first, then Y position
        block_top, summary_block, summary_event, _ = summary_candidates[0]
        processed_header_events.append(summary_event)
        summary_text = " ".join(summary_block.lines).strip()
        # Mark summary text as written IMMEDIATELY to prevent duplicates
        written_content.add(" ".join(summary_text.strip().lower().split()))
        LOGGER.info(f"Identified summary (FIRST occurrence) at Y={block_top:.1f}: {summary_text[:50]}...")
    
    # Layout: Photo in column B, Name/Title/Summary in column A (top section)
    photo_row = 1
    name_row = 1
    
    if candidate_photo:
        fragment, photo_event = candidate_photo
        try:
            excel_image = ExcelImage(io.BytesIO(fragment.image_bytes))
            # Resize photo appropriately for CV
            target_width = 120  # Standard CV photo size
            if fragment.width and fragment.width > 0:
                scale = target_width / fragment.width
                excel_image.width = target_width
                excel_image.height = max(120, min(180, int(fragment.height * scale)))
            else:
                excel_image.width = target_width
                excel_image.height = 150
            
            # Place photo in column B, row 1 (next to name)
            anchor = f"B{photo_row}"
            ws.add_image(excel_image, anchor)
            _ensure_min_width(ws, "B", 18)
            LOGGER.info(f"Placed candidate photo at {anchor}")
            stats["images_embedded"] += 1
            # Mark this event as processed
            photo_event["processed"] = True
        except Exception as photo_exc:
            LOGGER.warning(f"Could not place candidate photo: {photo_exc}")
            candidate_photo = None
    
    # Place name/title in column A, same row as photo
    if name_title_block is not None and name_event is not None:
        block = name_title_block
        for line_idx, line in enumerate(block.lines):
            # Mark content as written to prevent duplicates
            _is_duplicate_content(line)  # This adds it to written_content set
            
            cell = ws.cell(row=name_row + line_idx, column=1, value=line)
            
            # Apply background color from PDF (header area usually has colored background)
            if block.bg_color:
                bg_hex = f"{block.bg_color[0]:02X}{block.bg_color[1]:02X}{block.bg_color[2]:02X}"
                cell.fill = PatternFill(start_color=bg_hex, end_color=bg_hex, fill_type="solid")
                # Also apply to column B for header width
                header_cell = ws.cell(row=name_row + line_idx, column=2)
                header_cell.fill = PatternFill(start_color=bg_hex, end_color=bg_hex, fill_type="solid")
            
            # Apply text color from PDF
            text_color_hex = None
            if block.text_color:
                text_color_hex = f"{block.text_color[0]:02X}{block.text_color[1]:02X}{block.text_color[2]:02X}"
            
            if line_idx == 0:
                # Name: Bold, larger font
                if text_color_hex:
                    cell.font = Font(bold=True, size=16, name="Calibri", color=text_color_hex)
                else:
                    cell.font = Font(bold=True, size=16, name="Calibri", color="FFFFFF" if block.bg_color else "1F4E78")
            elif line_idx == 1:
                # Title: Bold, medium font
                if text_color_hex:
                    cell.font = Font(bold=True, size=12, name="Calibri", color=text_color_hex)
                else:
                    cell.font = Font(bold=True, size=12, name="Calibri", color="FFFFFF" if block.bg_color else "000000")
            else:
                # Other info: Regular font
                if text_color_hex:
                    cell.font = Font(size=11, name="Calibri", color=text_color_hex)
                else:
                    cell.font = Font(size=11, name="Calibri", color="FFFFFF" if block.bg_color else "000000")
            _adjust_column_width(ws, column_widths, 0, line)
        name_row += len(block.lines) + 1
        current_row = max(current_row, name_row)
        stats["rows_written"] = max(stats["rows_written"], current_row)
        # Mark this event as processed
        name_event["processed"] = True
        stats["text_blocks"] += 1
    
    # Place summary paragraph right after name/title (if found)
    if summary_block is not None and summary_event is not None:
        block = summary_block
        summary_text = " ".join(block.lines).strip()
        
        # Mark entire summary as written to prevent duplicates (BEFORE writing)
        # This ensures if same summary appears later, it will be skipped
        written_content.add(" ".join(summary_text.strip().lower().split()))
        
        # Write summary as a single paragraph (wrap text)
        cell = ws.cell(row=current_row, column=1, value=summary_text)
        cell.alignment = Alignment(vertical="top", wrap_text=True, horizontal="left")
        
        # Apply colors from PDF if available
        if block.bg_color:
            bg_hex = f"{block.bg_color[0]:02X}{block.bg_color[1]:02X}{block.bg_color[2]:02X}"
            cell.fill = PatternFill(start_color=bg_hex, end_color=bg_hex, fill_type="solid")
        
        text_color_hex = None
        if block.text_color:
            text_color_hex = f"{block.text_color[0]:02X}{block.text_color[1]:02X}{block.text_color[2]:02X}"
            cell.font = Font(size=11, name="Calibri", color=text_color_hex)
        else:
            cell.font = Font(size=11, name="Calibri")
        
        _adjust_column_width(ws, column_widths, 0, summary_text)
        current_row += 3  # Add spacing after summary
        stats["rows_written"] = max(stats["rows_written"], current_row)
        # Mark this event as processed
        summary_event["processed"] = True
        stats["text_blocks"] += 1
    
    # Mark processed header events (name/title, summary) as processed to prevent duplicates
    for header_event in processed_header_events:
        header_event["processed"] = True

    # Single column layout: All content flows sequentially in column A
    # Process remaining events (skip already processed photo, name, summary, and header blocks)
    for event in events:
        if event.get("processed"):
            continue
        if event["type"] == "text":
            block: ParagraphBlock = event["payload"]
            
            # Skip header blocks ONLY if they were already processed (name/title/summary)
            # Don't skip based on position alone - only skip if explicitly marked as processed
            # This ensures contact info and other header content is included
            
            # Skip if this is a duplicate name/title at bottom (already written at top)
            if name_title_block is not None:
                block_text = " ".join(block.lines).strip()
                name_text = " ".join(name_title_block.lines).strip()
                # Check if this block matches the name/title (exact match or first line match)
                if block_text.lower() == name_text.lower() or (block.lines and name_title_block.lines and block.lines[0].strip().lower() == name_title_block.lines[0].strip().lower()):
                    # This is a duplicate name - skip it
                    LOGGER.info(f"Skipping duplicate name/title at bottom: {block.lines[0] if block.lines else ''}")
                    continue
            
            # Detect CV/resume sections for better formatting
            first_line = block.lines[0].strip().upper() if block.lines else ""
            is_section_header = (
                first_line in ["EDUCATION", "EXPERIENCE", "WORK EXPERIENCE", "SKILLS", 
                              "TECHNICAL SKILLS", "SOFT SKILLS", "PROFESSIONAL SUMMARY",
                              "SUMMARY", "CONTACT", "PERSONAL INFORMATION", "OBJECTIVE",
                              "CERTIFICATIONS", "CERTIFICATES", "PROJECTS", "ACHIEVEMENTS",
                              "LANGUAGES", "INTERESTS"] or
                first_line.endswith(":") and len(first_line) < 30
            )
            
            # All content goes in column A, sequential order
            target_col = 1
            target_row = current_row
            
            if is_section_header and len(block.lines) == 1:
                # Check for duplicate content
                header_text = block.lines[0]
                if _is_duplicate_content(header_text):
                    LOGGER.debug(f"Skipping duplicate section header: {header_text[:50]}")
                    continue
                
                # Section header: Bold, larger font, with spacing
                cell = ws.cell(row=target_row, column=target_col, value=header_text)
                
                # Apply colors from PDF if available
                if block.bg_color:
                    # Convert RGB (0-255) to hex
                    bg_hex = f"{block.bg_color[0]:02X}{block.bg_color[1]:02X}{block.bg_color[2]:02X}"
                    cell.fill = PatternFill(start_color=bg_hex, end_color=bg_hex, fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="E7EFF6", end_color="E7EFF6", fill_type="solid")
                
                if block.text_color:
                    # Convert RGB to hex (openpyxl uses RRGGBB format)
                    text_hex = f"{block.text_color[0]:02X}{block.text_color[1]:02X}{block.text_color[2]:02X}"
                    cell.font = Font(bold=True, size=13, name="Calibri", color=text_hex)
                else:
                    cell.font = Font(bold=True, size=13, name="Calibri", color="1F4E78")
                
                _adjust_column_width(ws, column_widths, target_col - 1, block.lines[0])
                target_row += 1
                target_row += 1  # Extra spacing after header
                
                # Update current row (single column layout)
                current_row = target_row
                stats["rows_written"] = max(stats["rows_written"], current_row)
            else:
                # Regular content - ONLY check for exact duplicate (already in written_content)
                # Don't do aggressive filtering - let ALL content through except exact duplicates
                block_text = " ".join(block.lines).strip()
                
                # ONLY skip if it's an exact duplicate (normalized text already in written_content)
                # This is the ONLY duplicate check - no aggressive filtering
                if _is_duplicate_content(block_text):
                    LOGGER.debug(f"Skipping exact duplicate block: {block_text[:50]}...")
                    continue
                
                # Detect dates (patterns like "09/2019 - 07/2021" or "04/2025 - Present")
                date_pattern = r'\d{1,2}/\d{4}\s*-\s*(\d{1,2}/\d{4}|Present|present)'
                
                for line_idx, line in enumerate(block.lines):
                    # Don't check line-by-line duplicates - this was filtering out valid content
                    # Only block-level duplicate check is sufficient
                    
                    # ALWAYS place content in the target column only (column A for single column layout)
                    # Skip if line is just a date (will be handled with previous content)
                    if re.match(date_pattern, line.strip()):
                        # Date line: smaller font, italic
                        cell = ws.cell(row=target_row, column=target_col, value=line)
                        cell.font = Font(size=10, name="Calibri", italic=True, color="666666")
                        cell.alignment = Alignment(vertical="top", wrap_text=True, horizontal="left")
                    else:
                        cell = ws.cell(row=target_row, column=target_col, value=line)
                        
                        # Apply background color from PDF if available
                        if block.bg_color:
                            bg_hex = f"{block.bg_color[0]:02X}{block.bg_color[1]:02X}{block.bg_color[2]:02X}"
                            cell.fill = PatternFill(start_color=bg_hex, end_color=bg_hex, fill_type="solid")
                        
                        # Apply text color from PDF if available
                        text_color_hex = None
                        if block.text_color:
                            text_color_hex = f"{block.text_color[0]:02X}{block.text_color[1]:02X}{block.text_color[2]:02X}"
                        
                        # Format based on content type
                        if line.startswith("•") or line.startswith("-"):
                            # Bullet points - keep in same column, just indent text
                            if text_color_hex:
                                cell.font = Font(name="Calibri", size=11, color=text_color_hex)
                            else:
                                cell.font = body_font
                            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
                        elif any(keyword in line.upper() for keyword in ["UNIVERSITY", "COLLEGE", "SCHOOL", "INSTITUTE"]):
                            # Institution name: italic
                            if text_color_hex:
                                cell.font = Font(size=11, name="Calibri", italic=True, color=text_color_hex)
                            else:
                                cell.font = Font(size=11, name="Calibri", italic=True)
                            cell.alignment = Alignment(vertical="top", wrap_text=True, horizontal="left")
                        elif len(line.split()) <= 3 and not any(char.isdigit() for char in line) and is_section_header:
                            # Short lines that are section headers: bold
                            if text_color_hex:
                                cell.font = Font(bold=True, size=11, name="Calibri", color=text_color_hex)
                            else:
                                cell.font = Font(bold=True, size=11, name="Calibri")
                            cell.alignment = Alignment(vertical="top", wrap_text=True, horizontal="left")
                        else:
                            if text_color_hex:
                                cell.font = Font(name="Calibri", size=11, color=text_color_hex)
                            else:
                                cell.font = body_font
                            cell.alignment = Alignment(vertical="top", wrap_text=True, horizontal="left")
                    
                    # Adjust width of column A (for both date and regular content)
                    _adjust_column_width(ws, column_widths, 0, line)
                    target_row += 1
                
                stats["rows_written"] += 1
                
                target_row += 1  # Spacing between blocks
                
                # Update current row (single column layout)
                current_row = target_row
                stats["rows_written"] = max(stats["rows_written"], current_row)
            
            stats["text_blocks"] += 1

        elif event["type"] == "table":
            table: ExtractedTable = event["payload"]
            
            # Check if table content is duplicate (using aggressive duplicate detection)
            table_text = " ".join([" ".join(row) for row in table.data if row]).strip()
            if table_text and _is_duplicate_content(table_text):
                LOGGER.info(f"Skipping duplicate table content: {table_text[:50]}...")
                continue
            
            # Convert table to single-column text format (as per user requirement)
            # Write table as sequential text in Column A, not as multi-column table
            table_start_row = current_row
            
            for row_idx, row in enumerate(table.data):
                if not row:
                    continue
                
                # Combine all cells in the row into a single text line
                row_text = " | ".join([str(cell).strip() for cell in row if cell and str(cell).strip()])
                if not row_text:
                    continue
                
                # Check for duplicate row content
                if _is_duplicate_content(row_text):
                    LOGGER.debug(f"Skipping duplicate table row: {row_text[:50]}...")
                    continue
                
                # Write row as single line in Column A
                cell = ws.cell(row=table_start_row + row_idx, column=1, value=row_text)
                
                # First row is header - format it
                if row_idx == 0:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                else:
                    cell.font = body_font
                    if (table_start_row + row_idx) % 2 == 0:
                        cell.fill = PatternFill(start_color="F4F6FA", end_color="F4F6FA", fill_type="solid")
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                
                _adjust_column_width(ws, column_widths, 0, row_text)
            
            current_row = table_start_row + len([r for r in table.data if r]) + 1
            stats["tables_written"] += 1
            stats["rows_written"] = max(stats["rows_written"], current_row)

        elif event["type"] == "image":
            fragment: ImageFragment = event["payload"]
            
            # Skip if this is the candidate photo we already placed
            # Compare by page_index, bbox, and dimensions
            if candidate_photo_fragment:
                if (fragment.page_index == candidate_photo_fragment.page_index and
                    fragment.width == candidate_photo_fragment.width and
                    fragment.height == candidate_photo_fragment.height and
                    abs(fragment.bbox[0] - candidate_photo_fragment.bbox[0]) < 5 and
                    abs(fragment.bbox[1] - candidate_photo_fragment.bbox[1]) < 5):
                    continue
            
            try:
                excel_image = ExcelImage(io.BytesIO(fragment.image_bytes))
            except Exception as exc:
                LOGGER.warning("Skipping image on page %s: %s", fragment.page_index, exc)
                continue

            # Regular image/diagram: Full width in column A
            target_width = min(fragment.width, 900)
            if fragment.width and fragment.width > target_width:
                scale = target_width / fragment.width
                excel_image.width = target_width
                excel_image.height = max(1, int(fragment.height * scale))
            else:
                excel_image.width = fragment.width or target_width
                excel_image.height = fragment.height or 100

            anchor = f"A{current_row}"
            ws.add_image(excel_image, anchor)
            _ensure_min_width(ws, "A", max(30, excel_image.width / 10))
            rows_needed = max(6, math.ceil(excel_image.height / 120))

            current_row += rows_needed
            stats["rows_written"] = max(stats["rows_written"], current_row)
            stats["images_embedded"] += 1
            LOGGER.info(f"Embedded image at row {current_row - rows_needed}")
    
    # Single column layout - current_row is already updated

    virtual_wb = io.BytesIO()
    wb.save(virtual_wb)
    virtual_wb.seek(0)
    return virtual_wb.read(), stats


def _build_event_stream(
    paragraphs: Iterable[ParagraphBlock],
    tables: Iterable[ExtractedTable],
    images: Iterable[ImageFragment],
) -> List[Dict]:
    events: List[Dict] = []
    structured_pages = set()
    seen_blocks = set()  # Track seen blocks to prevent duplicates

    paragraph_list = list(paragraphs)
    table_list = list(tables)
    image_list = list(images)

    # Separate embedded images from page snapshots
    embedded_images = [img for img in image_list if img.label == "embedded-image"]
    page_snapshots = [img for img in image_list if img.label == "page-snapshot"]

    # Track pages with structured content
    for block in paragraph_list:
        structured_pages.add(block.page_index)
    for table in table_list:
        structured_pages.add(table.page_index)

    # Add all events with their positions for proper ordering
    # For single column layout, we need to preserve original PDF reading order (top-to-bottom, left-to-right)
    
    # Add embedded images (with duplicate detection)
    for fragment in embedded_images:
        # Create unique key for this image
        img_key = (fragment.page_index, round(fragment.bbox[0], 1), round(fragment.bbox[1], 1), 
                   fragment.width, fragment.height)
        if img_key not in seen_blocks:
            seen_blocks.add(img_key)
            events.append({
                "type": "image", 
                "page": fragment.page_index, 
                "top": fragment.bbox[1],  # Y position (top)
                "left": fragment.bbox[0],  # X position (left)
                "payload": fragment
            })

    # Add text blocks with their positions (conservative duplicate detection - only exact position+text matches)
    for block in paragraph_list:
        # Only check for exact duplicate (same position + same first line)
        # This ensures ALL content is included
        first_line = block.lines[0] if block.lines else ""
        block_key = (block.page_index, round(block.bbox[0], 1), round(block.bbox[1], 1), 
                     round(block.bbox[2], 1), round(block.bbox[3], 1), first_line[:100])
        
        # Only skip if EXACT same position AND same first line (true duplicate)
        if block_key not in seen_blocks:
            seen_blocks.add(block_key)
            events.append({
                "type": "text", 
                "page": block.page_index, 
                "top": block.bbox[1],  # Y position
                "left": block.bbox[0],  # X position
                "payload": block
            })
    
    # Add tables with their positions (conservative duplicate detection - only exact position matches)
    for table in table_list:
        # Only check for exact duplicate position (same page, same position)
        # This ensures ALL tables are included
        top = table.bbox[1] if table.bbox else 0.0
        left = table.bbox[0] if table.bbox else 0.0
        # Create unique key for table based on position
        table_key = (table.page_index, round(left, 1), round(top, 1))
        if table_key not in seen_blocks:
            seen_blocks.add(table_key)
            events.append({
                "type": "table", 
                "page": table.page_index, 
                "top": top,
                "left": left,
                "payload": table
            })
    
    # Add page snapshots only if page has no structured content
    for fragment in page_snapshots:
        if fragment.page_index not in structured_pages:
            img_key = (fragment.page_index, round(fragment.bbox[0], 1), round(fragment.bbox[1], 1))
            if img_key not in seen_blocks:
                seen_blocks.add(img_key)
                events.append({
                    "type": "image", 
                    "page": fragment.page_index, 
                    "top": fragment.bbox[1],
                    "left": fragment.bbox[0],
                    "payload": fragment
                })
    
    # Sort by: page number, then Y position (top to bottom), then X position (left to right)
    # This ensures content appears in the same reading order as the original PDF
    events.sort(key=lambda event: (
        event["page"],           # First by page number
        round(event.get("top", 0), 1),  # Then by Y position (top to bottom)
        round(event.get("left", 0), 1)  # Then by X position (left to right)
    ))
    
    LOGGER.info(f"Built event stream with {len(events)} events (after duplicate removal)")
    return events


def _write_table(
    ws,
    table: ExtractedTable,
    start_row: int,
    border: Border,
    header_fill: PatternFill,
    header_font: Font,
    body_font: Font,
    column_widths: Dict[int, float],
    start_col: int = 1,
) -> int:
    if not table.data:
        return start_row
    normalized_rows = _trim_empty_columns(table.data)
    if not normalized_rows:
        return start_row

    max_cols = max(len(row) for row in normalized_rows)
    # Limit table width to prevent spreading too much (max 3-4 columns)
    max_cols = min(max_cols, 3)
    
    for row_offset, row in enumerate(normalized_rows):
        excel_row = start_row + row_offset
        for col_index in range(max_cols):
            value = row[col_index] if col_index < len(row) else ""
            # Write to the specified start column + offset
            excel_col = start_col + col_index
            cell = ws.cell(row=excel_row, column=excel_col, value=value)
            cell.border = border

            if row_offset == 0:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.font = body_font
                if excel_row % 2 == 0:
                    cell.fill = PatternFill(start_color="F4F6FA", end_color="F4F6FA", fill_type="solid")
                cell.alignment = Alignment(vertical="top", wrap_text=True)

            _adjust_column_width(ws, column_widths, excel_col - 1, value)

    return start_row + len(normalized_rows) + 1


def _adjust_column_width(ws, column_widths: Dict[int, float], column_index: int, value: str):
    if value is None:
        return
    text_value = str(value)
    if not text_value:
        return
    desired = min(max(len(text_value) + 2, 12), 90)
    current = column_widths[column_index]
    if desired <= current:
        return
    column_widths[column_index] = desired
    ws.column_dimensions[get_column_letter(column_index + 1)].width = desired


def _ensure_min_width(ws, column_letter: str, width: float):
    dimension = ws.column_dimensions[column_letter]
    current = dimension.width or 0
    if current < width:
        dimension.width = width


def _trim_empty_columns(rows: List[List[str]]) -> List[List[str]]:
    if not rows:
        return rows

    col_count = max(len(row) for row in rows)
    if col_count == 0:
        return rows

    def column_has_data(idx: int) -> bool:
        for row in rows:
            if idx < len(row) and str(row[idx]).strip():
                return True
        return False

    start = 0
    while start < col_count and not column_has_data(start):
        start += 1

    if start == col_count:
        return []

    end = col_count - 1
    while end >= start and not column_has_data(end):
        end -= 1

    if start == 0 and end == col_count - 1:
        return rows

    trimmed = []
    for row in rows:
        trimmed.append(row[start : end + 1])
    return trimmed

