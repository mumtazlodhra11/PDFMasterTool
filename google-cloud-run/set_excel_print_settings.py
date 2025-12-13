#!/usr/bin/env python3
"""
Python script to set Excel print settings using openpyxl
This ensures charts don't split across pages
"""
import sys
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def set_print_settings_for_charts(input_path: str, output_path: str):
    """Set aggressive print settings to prevent chart splitting"""
    try:
        wb = load_workbook(input_path)
        
        for sheet in wb.worksheets:
            # CRITICAL: Fit to 1 page wide, unlimited height
            sheet.page_setup.fitToWidth = 1
            sheet.page_setup.fitToHeight = 0  # 0 = fit all rows
            sheet.page_setup.fitToPage = True
            sheet.page_setup.scale = None  # Must be None for fitToPage
            
            # Remove all page breaks
            if hasattr(sheet, 'row_breaks'):
                sheet.row_breaks = []
            if hasattr(sheet, 'col_breaks'):
                sheet.col_breaks = []
            
            # Set print area
            if sheet.max_row > 0 and sheet.max_column > 0:
                max_col = get_column_letter(sheet.max_column)
                sheet.print_area = f'A1:{max_col}{sheet.max_row}'
            
            # High quality
            sheet.page_setup.horizontalDpi = 300
            sheet.page_setup.verticalDpi = 300
        
        wb.save(output_path)
        print(f"✅ Print settings applied: {output_path}")
        return True
    except Exception as e:
        print(f"❌ Error: {e}")
        return False

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python set_excel_print_settings.py <input.xlsx> <output.xlsx>")
        sys.exit(1)
    
    set_print_settings_for_charts(sys.argv[1], sys.argv[2])

