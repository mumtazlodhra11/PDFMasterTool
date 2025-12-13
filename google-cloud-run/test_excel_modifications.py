#!/usr/bin/env python3
"""
Test script to verify Excel file modifications are working correctly
This can be run locally to test the Excel modification logic
"""
import sys
import os
import tempfile
from openpyxl import load_workbook, Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter

def create_test_excel():
    """Create a test Excel file with a chart"""
    print("Creating test Excel file...")
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Chart"
    
    # Add data
    data = [
        ['Month', 'Product A', 'Product B', 'Product C'],
        ['Jan', 1200, 900, 600],
        ['Feb', 1500, 1100, 750],
        ['Mar', 1700, 1300, 900],
        ['Apr', 1600, 1250, 850],
        ['May', 1800, 1400, 950],
        ['Jun', 2000, 1500, 1100],
    ]
    
    for row in data:
        ws.append(row)
    
    # Create chart
    chart = LineChart()
    chart.title = "Half-Yearly Sales Trend"
    chart.style = 13
    chart.y_axis.title = 'Sales'
    chart.x_axis.title = 'Month'
    
    data_ref = Reference(ws, min_col=2, min_row=1, max_col=4, max_row=7)
    cats_ref = Reference(ws, min_col=1, min_row=2, max_row=7)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws.add_chart(chart, "E2")
    
    return wb

def apply_modifications(wb, input_path, output_path):
    """Apply the same modifications as the backend"""
    print("\n" + "="*60)
    print("APPLYING EXCEL MODIFICATIONS (Same as Backend)")
    print("="*60 + "\n")
    
    prevent_chart_split_bool = True
    fit_to_page_bool = True
    orientation = None
    page_size = "A4"
    
    for sheet in wb.worksheets:
        print(f"[XLSX->PDF] Processing sheet: {sheet.title}")
        
        # Check original settings
        print(f"\n  Original settings:")
        print(f"    fitToWidth: {sheet.page_setup.fitToWidth}")
        print(f"    fitToHeight: {sheet.page_setup.fitToHeight}")
        print(f"    fitToPage: {sheet.page_setup.fitToPage}")
        print(f"    orientation: {sheet.page_setup.orientation}")
        print(f"    row_breaks: {len(sheet.row_breaks) if hasattr(sheet, 'row_breaks') else 'N/A'}")
        print(f"    col_breaks: {len(sheet.col_breaks) if hasattr(sheet, 'col_breaks') else 'N/A'}")
        
        # Apply modifications
        if fit_to_page_bool or prevent_chart_split_bool:
            sheet.page_setup.fitToWidth = 1
            sheet.page_setup.fitToHeight = 999
            sheet.page_setup.fitToPage = True
            sheet.page_setup.scale = None
            sheet.page_setup.blackAndWhite = False
            sheet.page_setup.draft = False
            try:
                sheet.page_setup.pageOrder = 'downThenOver'
            except:
                pass
            print(f"\n  ✅✅✅ AGGRESSIVE: fitToWidth=1, fitToHeight=999, fitToPage=True, scale=None")
        
        # Remove page breaks
        try:
            if hasattr(sheet, 'row_breaks'):
                sheet.row_breaks = []
            if hasattr(sheet, 'col_breaks'):
                sheet.col_breaks = []
            print(f"  ✅✅✅ Removed ALL page breaks")
        except Exception as e:
            print(f"  ⚠️ Warning: Could not remove page breaks: {e}")
        
        # Set print area
        if sheet.max_row > 0 and sheet.max_column > 0:
            max_col_letter = get_column_letter(sheet.max_column)
            print_area = f'A1:{max_col_letter}{sheet.max_row}'
            sheet.print_area = print_area
            print(f"  ✅ Set print area: {print_area}")
        
        # Force landscape
        if prevent_chart_split_bool:
            sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
            print(f"  ✅✅✅ FORCED: Landscape orientation")
        
        # Set page size
        if page_size:
            size_map = {
                "A4": 9,
                "Letter": 1,
                "Legal": 5,
                "A3": 8,
                "Tabloid": 3,
            }
            if page_size in size_map:
                sheet.page_setup.paperSize = size_map[page_size]
                print(f"  ✅ Set paper size: {page_size}")
        
        # Minimal margins
        if prevent_chart_split_bool:
            sheet.page_margins.top = 0.3
            sheet.page_margins.bottom = 0.3
            sheet.page_margins.left = 0.3
            sheet.page_margins.right = 0.3
            sheet.page_margins.header = 0.2
            sheet.page_margins.footer = 0.2
            print(f"  ✅✅✅ FORCED: Minimal margins (0.3\")")
        
        # High quality
        sheet.page_setup.horizontalDpi = 300
        sheet.page_setup.verticalDpi = 300
        print(f"  ✅ Set high quality (300 DPI)")
        
        # Verify modifications
        print(f"\n  Modified settings:")
        print(f"    fitToWidth: {sheet.page_setup.fitToWidth}")
        print(f"    fitToHeight: {sheet.page_setup.fitToHeight}")
        print(f"    fitToPage: {sheet.page_setup.fitToPage}")
        print(f"    orientation: {sheet.page_setup.orientation}")
        print(f"    row_breaks: {len(sheet.row_breaks) if hasattr(sheet, 'row_breaks') else 'N/A'}")
        print(f"    col_breaks: {len(sheet.col_breaks) if hasattr(sheet, 'col_breaks') else 'N/A'}")
        print(f"    margins: top={sheet.page_margins.top}, bottom={sheet.page_margins.bottom}, left={sheet.page_margins.left}, right={sheet.page_margins.right}")
    
    # Save modified file
    wb.save(output_path)
    print(f"\n✅✅✅ Excel file modified and saved: {output_path}")
    print(f"   File size: {os.path.getsize(output_path)} bytes")
    
    return True

def verify_modifications(file_path):
    """Verify that modifications were saved correctly"""
    print("\n" + "="*60)
    print("VERIFYING MODIFICATIONS")
    print("="*60 + "\n")
    
    wb = load_workbook(file_path)
    for sheet in wb.worksheets:
        print(f"Sheet: {sheet.title}")
        print(f"  fitToWidth: {sheet.page_setup.fitToWidth} (expected: 1)")
        print(f"  fitToHeight: {sheet.page_setup.fitToHeight} (expected: 999)")
        print(f"  fitToPage: {sheet.page_setup.fitToPage} (expected: True)")
        print(f"  orientation: {sheet.page_setup.orientation} (expected: landscape)")
        print(f"  row_breaks: {len(sheet.row_breaks) if hasattr(sheet, 'row_breaks') else 'N/A'} (expected: 0)")
        print(f"  col_breaks: {len(sheet.col_breaks) if hasattr(sheet, 'col_breaks') else 'N/A'} (expected: 0)")
        print(f"  margins: top={sheet.page_margins.top}, bottom={sheet.page_margins.bottom}")
        
        # Check if modifications are correct
        all_correct = (
            sheet.page_setup.fitToWidth == 1 and
            sheet.page_setup.fitToHeight == 999 and
            sheet.page_setup.fitToPage == True and
            sheet.page_setup.orientation == sheet.ORIENTATION_LANDSCAPE and
            len(sheet.row_breaks) == 0 and
            len(sheet.col_breaks) == 0
        )
        
        if all_correct:
            print(f"  ✅✅✅ ALL MODIFICATIONS CORRECT!")
        else:
            print(f"  ❌❌❌ SOME MODIFICATIONS MISSING!")
    
    return all_correct

if __name__ == "__main__":
    print("="*60)
    print("EXCEL MODIFICATION TEST SCRIPT")
    print("="*60)
    
    with tempfile.TemporaryDirectory() as temp_dir:
        # Create original Excel file
        wb = create_test_excel()
        original_path = os.path.join(temp_dir, "original.xlsx")
        wb.save(original_path)
        print(f"✅ Original Excel file created: {original_path}")
        
        # Apply modifications
        modified_path = os.path.join(temp_dir, "modified.xlsx")
        apply_modifications(wb, original_path, modified_path)
        
        # Verify modifications
        verify_modifications(modified_path)
        
        print("\n" + "="*60)
        print("TEST COMPLETE")
        print("="*60)
        print("\n⚠️  IMPORTANT:")
        print("   If modifications are correct, the issue is LibreOffice limitation.")
        print("   LibreOffice doesn't respect Excel print settings for charts.")
        print("   Charts are floating objects and LibreOffice places them based on")
        print("   sheet position, not print settings.\n")
        print(f"📁 Test files saved in: {temp_dir}")
        print(f"   - Original: {original_path}")
        print(f"   - Modified: {modified_path}\n")

