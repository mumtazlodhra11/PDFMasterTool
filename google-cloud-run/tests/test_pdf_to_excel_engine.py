import base64
import io
from typing import List

import fitz
import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from PIL import Image as PILImage

from app import app
from pdf_to_excel_engine import convert_pdf_to_excel
import pdf_to_excel_engine.table_extractor as table_mod


TABLE_LIB_AVAILABLE = bool(getattr(table_mod, "camelot", None) or getattr(table_mod, "tabula_read_pdf", None))
client = TestClient(app)


def _create_table_pdf(include_diagram: bool = False) -> bytes:
    doc = fitz.open()
    page = doc.new_page(width=595, height=842)

    top = 120
    left = 60
    row_height = 40
    col_width = 140
    rows = 4
    cols = 3

    # Draw table grid
    for r in range(rows + 1):
        y = top + r * row_height
        page.draw_line(fitz.Point(left, y), fitz.Point(left + cols * col_width, y), width=1)
    for c in range(cols + 1):
        x = left + c * col_width
        page.draw_line(fitz.Point(x, top), fitz.Point(x, top + rows * row_height), width=1)

    headers = ["Provider", "IP Range", "Gateway"]
    data = [
        ["ISP-1", "10.0.0.0/24", "gw1.example.net"],
        ["ISP-2", "10.0.1.0/24", "gw2.example.net"],
        ["ISP-3", "10.0.2.0/24", "gw3.example.net"],
    ]

    for idx, header in enumerate(headers):
        rect = fitz.Rect(left + idx * col_width + 4, top + 4, left + (idx + 1) * col_width - 4, top + row_height - 4)
        page.insert_textbox(rect, header, fontsize=12, align=fitz.TEXT_ALIGN_CENTER)

    for row_idx, row in enumerate(data, start=1):
        for col_idx, cell in enumerate(row):
            rect = fitz.Rect(
                left + col_idx * col_width + 4,
                top + row_idx * row_height + 4,
                left + (col_idx + 1) * col_width - 4,
                top + (row_idx + 1) * row_height - 4,
            )
            page.insert_textbox(rect, cell, fontsize=11)

    if include_diagram:
        diagram = PILImage.new("RGB", (200, 120), color=(45, 125, 210))
        buf = io.BytesIO()
        diagram.save(buf, format="PNG")
        page.insert_image(fitz.Rect(320, 350, 520, 470), stream=buf.getvalue())

    pdf_bytes = doc.tobytes()
    doc.close()
    return pdf_bytes


def _create_text_only_pdf() -> bytes:
    doc = fitz.open()
    page = doc.new_page(width=595, height=842)
    paragraphs: List[str] = [
        "Network migration plan:",
        "1. Prepare redundant ISP links.",
        "2. Schedule maintenance window.",
        "3. Notify stakeholders 24 hours in advance.",
        "4. Validate monitoring dashboards.",
    ]
    text = "\n".join(paragraphs)
    page.insert_text((60, 120), text, fontsize=12, leading=18)
    pdf_bytes = doc.tobytes()
    doc.close()
    return pdf_bytes


@pytest.mark.skipif(not TABLE_LIB_AVAILABLE, reason="Table extraction libraries unavailable")
def test_convert_pdf_tables_only():
    pdf_bytes = _create_table_pdf()
    xlsx_bytes, stats = convert_pdf_to_excel(pdf_bytes)
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active
    values = [cell.value for cell in ws[3][0:3]]
    assert "ISP-1" in values
    assert stats["tables_detected"] >= 1


@pytest.mark.skipif(not TABLE_LIB_AVAILABLE, reason="Table extraction libraries unavailable")
def test_convert_pdf_tables_and_diagrams():
    pdf_bytes = _create_table_pdf(include_diagram=True)
    _, stats = convert_pdf_to_excel(pdf_bytes)
    assert stats["tables_detected"] >= 1
    assert stats["images_extracted"] >= 1


def test_convert_pdf_text_only():
    pdf_bytes = _create_text_only_pdf()
    xlsx_bytes, stats = convert_pdf_to_excel(pdf_bytes)
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active
    first_column_values = [ws.cell(row=row, column=1).value for row in range(1, ws.max_row + 1)]
    assert any("Network migration plan" in (value or "") for value in first_column_values)
    assert stats["text_blocks"] >= 1


def test_pdf_to_excel_endpoint_returns_xlsx():
    pdf_bytes = _create_text_only_pdf()
    response = client.post(
        "/convert/pdf-to-excel",
        files={"file": ("plan.pdf", pdf_bytes, "application/pdf")},
    )
    assert response.status_code == 200
    payload = response.json()
    assert payload["success"] is True
    file_bytes = base64.b64decode(payload["file"])
    wb = load_workbook(io.BytesIO(file_bytes))
    assert wb.active.max_row >= 2

