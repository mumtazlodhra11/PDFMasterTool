#!/usr/bin/env python3
"""
Test script to verify Excel chart split prevention
This script tests if LibreOffice respects Excel print settings for charts
"""
import sys
import os
import tempfile
from openpyxl import load_workbook, Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter
import subprocess

def create_test_excel_with_chart():
    """Create a test Excel file with a wide chart that might split"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Chart"
    
    # Add some data
    data = [
        ['Month', 'Product A', 'Product B', 'Product C'],
        ['Jan', 1200, 900, 600],
        ['Feb', 1500, 1100, 750],
        ['Mar', 1700, 1300, 900],
        ['Apr', 1600, 1250, 850],
        ['May', 1800, 1400, 950],
        ['Jun', 2000, 1500, 1100],
    ]
    
    for row in data:
        ws.append(row)
    
    # Create a wide chart
    chart = LineChart()
    chart.title = "Half-Yearly Sales Trend"
    chart.style = 13
    chart.y_axis.title = 'Sales'
    chart.x_axis.title = 'Month'
    
    data_ref = Reference(ws, min_col=2, min_row=1, max_col=4, max_row=7)
    cats_ref = Reference(ws, min_col=1, min_row=2, max_row=7)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    
    # Position chart to be wide (might cause splitting)
    ws.add_chart(chart, "E2")
    
    # Apply print settings to prevent splitting
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.fitToPage = True
    ws.page_setup.scale = None
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = 9  # A4
    
    # Remove page breaks
    ws.row_breaks = []
    ws.col_breaks = []
    
    # Set minimal margins
    ws.page_margins.top = 0.3
    ws.page_margins.bottom = 0.3
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    
    # Set print area
    ws.print_area = f'A1:{get_column_letter(ws.max_column)}{ws.max_row}'
    
    return wb

def test_conversion(input_path, output_path):
    """Test LibreOffice conversion"""
    print(f"\n{'='*60}")
    print("Testing Excel to PDF conversion with chart split prevention")
    print(f"{'='*60}\n")
    
    print(f"Input file: {input_path}")
    print(f"Output file: {output_path}\n")
    
    # Convert using LibreOffice
    cmd = [
        'soffice',
        '--headless',
        '--invisible',
        '--nologo',
        '--nodefault',
        '--convert-to', 'pdf:"calc_pdf_Export:{IsLandscape=true,Quality=100}"',
        '--outdir', os.path.dirname(output_path),
        input_path
    ]
    
    print(f"LibreOffice command: {' '.join(cmd)}\n")
    
    env = os.environ.copy()
    env['HOME'] = '/tmp'
    env['SAL_USE_VCLPLUGIN'] = 'headless'
    
    try:
        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            timeout=60,
            env=env
        )
        
        print(f"Return code: {result.returncode}")
        if result.stdout:
            print(f"Stdout: {result.stdout}")
        if result.stderr:
            print(f"Stderr: {result.stderr}")
        
        if os.path.exists(output_path):
            size = os.path.getsize(output_path)
            print(f"\n✅ PDF created: {output_path} ({size} bytes)")
            
            # Check PDF pages (would need PyPDF2 or similar)
            print("\n⚠️  NOTE: To check if charts are split, open the PDF and verify:")
            print("   1. Charts should be on single pages")
            print("   2. No chart should be split across pages")
            print("   3. Landscape orientation should be used")
            
            return True
        else:
            print(f"\n❌ PDF not created at: {output_path}")
            return False
            
    except Exception as e:
        print(f"\n❌ Error: {e}")
        return False

if __name__ == "__main__":
    with tempfile.TemporaryDirectory() as temp_dir:
        # Create test Excel file
        print("Creating test Excel file with chart...")
        wb = create_test_excel_with_chart()
        
        input_path = os.path.join(temp_dir, "test_chart.xlsx")
        wb.save(input_path)
        print(f"✅ Test Excel file created: {input_path}")
        
        # Verify print settings
        wb_check = load_workbook(input_path)
        ws_check = wb_check.active
        print(f"\nPrint settings in Excel file:")
        print(f"  fitToWidth: {ws_check.page_setup.fitToWidth}")
        print(f"  fitToHeight: {ws_check.page_setup.fitToHeight}")
        print(f"  fitToPage: {ws_check.page_setup.fitToPage}")
        print(f"  orientation: {ws_check.page_setup.orientation}")
        print(f"  paperSize: {ws_check.page_setup.paperSize}")
        print(f"  row_breaks: {len(ws_check.row_breaks)}")
        print(f"  col_breaks: {len(ws_check.col_breaks)}")
        
        # Test conversion
        output_path = os.path.join(temp_dir, "test_chart.pdf")
        test_conversion(input_path, output_path)
        
        print(f"\n{'='*60}")
        print("Test complete!")
        print(f"{'='*60}\n")
        print("⚠️  IMPORTANT: LibreOffice may not respect Excel print settings for charts.")
        print("   Charts are floating objects and LibreOffice places them based on")
        print("   sheet position, not print settings. This is a known limitation.\n")

